"""
Script para crear archivo edge_cases_composite_piers.xlsx con piers de formas L, T, C.

Este archivo contiene todas las tablas necesarias para probar la importación
de piers compuestos desde ETABS:
- Wall Object Connectivity: define cada muro como rectángulo con 4 puntos
- Point Object Connectivity: coordenadas X, Y, Z de todos los puntos
- Area Assigns - Pier Labels: agrupa múltiples muros en un Pier
- Pier Section Properties: propiedades básicas de los piers
- Pier Forces: fuerzas para análisis
"""
import openpyxl
from openpyxl import Workbook

wb = Workbook()

# ============================================================================
# Helper: crear puntos para un rectángulo dado centro, largo, espesor y ángulo
# ============================================================================
def create_wall_points(cx, cy, length, thickness, angle_deg=0):
    """
    Crea 4 puntos para un muro rectangular.
    - cx, cy: centro del muro
    - length: largo del muro (dirección principal)
    - thickness: espesor del muro
    - angle_deg: 0 = horizontal (largo en X), 90 = vertical (largo en Y)
    """
    import math
    angle = math.radians(angle_deg)

    # Vectores unitarios
    dx = math.cos(angle)
    dy = math.sin(angle)
    # Perpendicular
    px = -dy
    py = dx

    half_L = length / 2
    half_T = thickness / 2

    # 4 esquinas: partiendo de esquina inferior izquierda, antihorario
    p1 = (cx - dx * half_L - px * half_T, cy - dy * half_L - py * half_T)  # bottom-left
    p2 = (cx + dx * half_L - px * half_T, cy + dy * half_L - py * half_T)  # bottom-right
    p3 = (cx + dx * half_L + px * half_T, cy + dy * half_L + py * half_T)  # top-right
    p4 = (cx - dx * half_L + px * half_T, cy - dy * half_L + py * half_T)  # top-left

    return [p1, p2, p3, p4]

# ============================================================================
# Definir geometrías de piers compuestos
# ============================================================================

# Almacenar todos los puntos y muros
all_points = {}  # point_id -> (x, y, z)
all_walls = []   # [(wall_id, story, wall_bay, pt1, pt2, pt3, pt4)]
all_pier_assigns = []  # [(story, wall_label, wall_id, pier_name)]

point_id_counter = 1
wall_id_counter = 1
z_level = 3.0  # Altura del piso

def add_point(x, y, z=z_level):
    """Agrega un punto y retorna su ID."""
    global point_id_counter
    # Redondear para evitar duplicados por precisión flotante
    x, y, z = round(x, 4), round(y, 4), round(z, 4)

    # Buscar si ya existe
    for pid, (px, py, pz) in all_points.items():
        if abs(px - x) < 0.001 and abs(py - y) < 0.001 and abs(pz - z) < 0.001:
            return pid

    pid = point_id_counter
    all_points[pid] = (x, y, z)
    point_id_counter += 1
    return pid

def add_wall(story, pier_name, cx, cy, length, thickness, angle_deg=0):
    """Agrega un muro rectangular y lo asigna a un pier."""
    global wall_id_counter

    points = create_wall_points(cx, cy, length, thickness, angle_deg)
    pt_ids = [add_point(p[0], p[1]) for p in points]

    wall_id = wall_id_counter
    wall_bay = f"W{wall_id}"

    all_walls.append((wall_id, story, wall_bay, pt_ids[0], pt_ids[1], pt_ids[2], pt_ids[3]))
    all_pier_assigns.append((story, wall_bay, wall_id, pier_name))

    wall_id_counter += 1
    return wall_id

# ============================================================================
# PIER PL1: Forma L (1200mm x 200mm horizontal + 800mm x 200mm vertical)
# ============================================================================
story = "Piso1"

# Ala horizontal: centro en (0.6, 0), largo=1.2m, espesor=0.2m
add_wall(story, "PL1", cx=0.6, cy=0, length=1.2, thickness=0.2, angle_deg=0)
# Alma vertical: conectada al final del ala, centro en (1.2, 0.4), largo=0.8m, espesor=0.2m
add_wall(story, "PL1", cx=1.2, cy=0.4, length=0.8, thickness=0.2, angle_deg=90)

# ============================================================================
# PIER PL2: Forma L invertida (más grande)
# ============================================================================
# Ala horizontal: 1500mm x 250mm
add_wall(story, "PL2", cx=4.75, cy=0, length=1.5, thickness=0.25, angle_deg=0)
# Alma vertical: 1000mm x 250mm, conectada al inicio
add_wall(story, "PL2", cx=4.0, cy=0.5, length=1.0, thickness=0.25, angle_deg=90)

# ============================================================================
# PIER PT1: Forma T (ala horizontal 1400mm x 200mm, alma vertical 1000mm x 200mm)
# ============================================================================
# Ala horizontal centrada
add_wall(story, "PT1", cx=8.0, cy=0, length=1.4, thickness=0.2, angle_deg=0)
# Alma vertical desde el centro del ala hacia abajo
add_wall(story, "PT1", cx=8.0, cy=-0.5, length=1.0, thickness=0.2, angle_deg=90)

# ============================================================================
# PIER PT2: Forma T más ancha
# ============================================================================
# Ala horizontal: 2000mm x 200mm
add_wall(story, "PT2", cx=12.0, cy=0, length=2.0, thickness=0.2, angle_deg=0)
# Alma vertical: 1200mm x 250mm
add_wall(story, "PT2", cx=12.0, cy=-0.6, length=1.2, thickness=0.25, angle_deg=90)

# ============================================================================
# PIER PC1: Forma C (2 alas horizontales + alma vertical)
# ============================================================================
# Alma vertical central: 1500mm x 200mm
add_wall(story, "PC1", cx=16.0, cy=0, length=1.5, thickness=0.2, angle_deg=90)
# Ala superior: 800mm x 200mm
add_wall(story, "PC1", cx=16.4, cy=0.75, length=0.8, thickness=0.2, angle_deg=0)
# Ala inferior: 800mm x 200mm
add_wall(story, "PC1", cx=16.4, cy=-0.75, length=0.8, thickness=0.2, angle_deg=0)

# ============================================================================
# PIER PC2: Forma C más grande
# ============================================================================
# Alma vertical: 2000mm x 250mm
add_wall(story, "PC2", cx=20.0, cy=0, length=2.0, thickness=0.25, angle_deg=90)
# Ala superior: 1000mm x 250mm
add_wall(story, "PC2", cx=20.5, cy=1.0, length=1.0, thickness=0.25, angle_deg=0)
# Ala inferior: 1000mm x 250mm
add_wall(story, "PC2", cx=20.5, cy=-1.0, length=1.0, thickness=0.25, angle_deg=0)

# ============================================================================
# PIER P1: Pier rectangular simple (para comparación)
# ============================================================================
add_wall(story, "P1", cx=24.0, cy=0, length=1.0, thickness=0.2, angle_deg=0)

print(f"Total puntos: {len(all_points)}")
print(f"Total muros: {len(all_walls)}")
print(f"Total asignaciones: {len(all_pier_assigns)}")

# ============================================================================
# Sheet 1: Point Object Connectivity
# ============================================================================
ws1 = wb.active
ws1.title = 'Point Object Connectivity'
ws1['A1'] = 'TABLE:  Point Object Connectivity'

headers1 = ['UniqueName', 'Is Auto Point', 'Story', 'PointBay', 'IsSpecial', 'X', 'Y', 'Z', 'GUID']
for col, h in enumerate(headers1, 1):
    ws1.cell(row=2, column=col, value=h)

units1 = ['', '', '', '', '', 'm', 'm', 'm', '']
for col, u in enumerate(units1, 1):
    ws1.cell(row=3, column=col, value=u)

for row_idx, (pid, (x, y, z)) in enumerate(sorted(all_points.items()), 4):
    ws1.cell(row=row_idx, column=1, value=pid)
    ws1.cell(row=row_idx, column=2, value='No')
    ws1.cell(row=row_idx, column=3, value='Piso1')
    ws1.cell(row=row_idx, column=4, value=pid)
    ws1.cell(row=row_idx, column=5, value='No')
    ws1.cell(row=row_idx, column=6, value=x)
    ws1.cell(row=row_idx, column=7, value=y)
    ws1.cell(row=row_idx, column=8, value=z)
    ws1.cell(row=row_idx, column=9, value=f'guid-{pid}')

# ============================================================================
# Sheet 2: Wall Object Connectivity
# ============================================================================
ws2 = wb.create_sheet('Wall Object Connectivity')
ws2['A1'] = 'TABLE:  Wall Object Connectivity'

headers2 = ['UniqueName', 'Story', 'WallBay', 'UniquePt1', 'UniquePt2', 'UniquePt3', 'UniquePt4', 'Perimeter', 'Area', 'GUID']
for col, h in enumerate(headers2, 1):
    ws2.cell(row=2, column=col, value=h)

units2 = ['', '', '', '', '', '', '', 'm', 'm²', '']
for col, u in enumerate(units2, 1):
    ws2.cell(row=3, column=col, value=u)

for row_idx, (wall_id, story, wall_bay, pt1, pt2, pt3, pt4) in enumerate(all_walls, 4):
    ws2.cell(row=row_idx, column=1, value=wall_id)
    ws2.cell(row=row_idx, column=2, value=story)
    ws2.cell(row=row_idx, column=3, value=wall_bay)
    ws2.cell(row=row_idx, column=4, value=pt1)
    ws2.cell(row=row_idx, column=5, value=pt2)
    ws2.cell(row=row_idx, column=6, value=pt3)
    ws2.cell(row=row_idx, column=7, value=pt4)
    # Calcular perímetro y área aproximados
    ws2.cell(row=row_idx, column=8, value=3.0)  # Aproximado
    ws2.cell(row=row_idx, column=9, value=0.5)  # Aproximado
    ws2.cell(row=row_idx, column=10, value=f'wall-guid-{wall_id}')

# ============================================================================
# Sheet 3: Area Assigns - Pier Labels
# ============================================================================
ws3 = wb.create_sheet('Area Assigns - Pier Labels')
ws3['A1'] = 'TABLE:  Area Assignments - Pier Labels'

headers3 = ['Story', 'Label', 'UniqueName', 'Pier Name']
for col, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=col, value=h)

for row_idx, (story, wall_label, wall_id, pier_name) in enumerate(all_pier_assigns, 3):
    ws3.cell(row=row_idx, column=1, value=story)
    ws3.cell(row=row_idx, column=2, value=wall_label)
    ws3.cell(row=row_idx, column=3, value=wall_id)
    ws3.cell(row=row_idx, column=4, value=pier_name)

# ============================================================================
# Sheet 4: Pier Section Properties
# ============================================================================
ws4 = wb.create_sheet('Pier Section Properties')
ws4['A1'] = 'TABLE:  Pier Section Properties'

headers4 = ['Story', 'Pier', 'AxisAngle', '# Area Objects', '# Line Objects',
            'Width Bottom', 'Thickness Bottom', 'Width Top', 'Thickness Top', 'Material']
for col, h in enumerate(headers4, 1):
    ws4.cell(row=2, column=col, value=h)

units4 = ['', '', 'deg', '', '', 'm', 'm', 'm', 'm', '']
for col, u in enumerate(units4, 1):
    ws4.cell(row=3, column=col, value=u)

# Propiedades de cada pier (valores aproximados, serán recalculados del composite)
pier_props = [
    # Story, Pier, Angle, AreaObj, LineObj, WBottom, TBottom, WTop, TTop, Material
    ('Piso1', 'PL1', 0, 2, 0, 1.2, 0.2, 1.2, 0.2, '3000Psi'),    # L-shape
    ('Piso1', 'PL2', 0, 2, 0, 1.5, 0.25, 1.5, 0.25, '3000Psi'),  # L-shape invertida
    ('Piso1', 'PT1', 0, 2, 0, 1.4, 0.2, 1.4, 0.2, '3000Psi'),    # T-shape
    ('Piso1', 'PT2', 0, 2, 0, 2.0, 0.25, 2.0, 0.25, '3000Psi'),  # T-shape grande
    ('Piso1', 'PC1', 0, 3, 0, 1.5, 0.2, 1.5, 0.2, '3000Psi'),    # C-shape
    ('Piso1', 'PC2', 0, 3, 0, 2.0, 0.25, 2.0, 0.25, '3000Psi'),  # C-shape grande
    ('Piso1', 'P1', 0, 1, 0, 1.0, 0.2, 1.0, 0.2, '3000Psi'),     # Rectangular
]
for row_idx, pier in enumerate(pier_props, 4):
    for col, val in enumerate(pier, 1):
        ws4.cell(row=row_idx, column=col, value=val)

# ============================================================================
# Sheet 5: Pier Forces
# ============================================================================
ws5 = wb.create_sheet('Pier Forces')
ws5['A1'] = 'TABLE:  Pier Forces'

headers5 = ['Story', 'Pier', 'Output Case', 'Case Type', 'Step Type', 'Location',
            'P', 'V2', 'V3', 'T', 'M2', 'M3']
for col, h in enumerate(headers5, 1):
    ws5.cell(row=2, column=col, value=h)

units5 = ['', '', '', '', '', '', 'tonf', 'tonf', 'tonf', 'tonf-m', 'tonf-m', 'tonf-m']
for col, u in enumerate(units5, 1):
    ws5.cell(row=3, column=col, value=u)

# Fuerzas para cada pier con diferentes ángulos de momento
# Cada combinación tiene diferente proporción M2/M3 para probar interpolación
pier_forces = []

# Datos base por pier: (P, V2, V3, T)
pier_base = {
    'PL1': (-40, 10, 2, 0.5),   # L pequeño
    'PL2': (-60, 15, 3, 0.8),   # L grande
    'PT1': (-50, 12, 2.5, 0.6), # T pequeño
    'PT2': (-80, 20, 4, 1.0),   # T grande
    'PC1': (-70, 18, 3.5, 0.9), # C pequeño
    'PC2': (-100, 25, 5, 1.2),  # C grande
    'P1':  (-35, 8, 1.5, 0.4),  # Rectangular
}

# Combinaciones con diferentes ángulos de momento
# Combo, Type, Step, Location, M2_factor, M3_factor, P_sign (1=compresión, -1=tracción)
# Diseñados para dar diferentes ángulos: 0°, 14°, 30°, 45°, 60°, 75°, 90°
combos_with_angles = [
    # === COMPRESIÓN (P negativo) ===
    # M3 dominante (ángulo cerca de 0°)
    ('1.2D+1.6L', 'Combination', '', 'Top', 1.0, 15.0, 1),      # ~3.8° (M3 >> M2)
    ('1.2D+1.6L', 'Combination', '', 'Bottom', 0.5, 10.0, 1),   # ~2.9°

    # Sismo X - M3 dominante pero con algo de M2 (~15°)
    ('1.4X+1.2D+1.0L', 'Combination', 'Max', 'Top', 4.0, 15.0, 1),     # ~14.9°
    ('1.4X+1.2D+1.0L', 'Combination', 'Max', 'Bottom', 3.0, 12.0, 1),  # ~14.0°

    # Sismo Y - M2 y M3 similares (~45°)
    ('1.4Y+1.2D+1.0L', 'Combination', 'Max', 'Top', 12.0, 12.0, 1),    # 45°
    ('1.4Y+1.2D+1.0L', 'Combination', 'Max', 'Bottom', 10.0, 10.0, 1), # 45°

    # Sismo diagonal (~30°)
    ('1.2D+1.0E_30', 'Combination', 'Max', 'Top', 8.0, 14.0, 1),       # ~29.7°
    ('1.2D+1.0E_30', 'Combination', 'Max', 'Bottom', 6.0, 10.0, 1),    # ~31.0°

    # Sismo diagonal (~60°)
    ('1.2D+1.0E_60', 'Combination', 'Max', 'Top', 14.0, 8.0, 1),       # ~60.3°
    ('1.2D+1.0E_60', 'Combination', 'Max', 'Bottom', 12.0, 7.0, 1),    # ~59.7°

    # M2 dominante (~75°)
    ('1.4Y_dom+1.2D', 'Combination', 'Max', 'Top', 18.0, 5.0, 1),      # ~74.5°
    ('1.4Y_dom+1.2D', 'Combination', 'Max', 'Bottom', 15.0, 4.0, 1),   # ~75.1°

    # M2 casi puro (~85°)
    ('1.4Y_pure+1.2D', 'Combination', 'Max', 'Top', 20.0, 2.0, 1),     # ~84.3°
    ('1.4Y_pure+1.2D', 'Combination', 'Max', 'Bottom', 18.0, 1.5, 1),  # ~85.2°

    # === TRACCIÓN (P positivo) ===
    # Tracción con M3 dominante
    ('0.9D-1.4X', 'Combination', 'Min', 'Top', 2.0, 10.0, -1),         # ~11.3°, P tracción
    ('0.9D-1.4X', 'Combination', 'Min', 'Bottom', 1.5, 8.0, -1),       # ~10.6°, P tracción

    # Tracción con ángulo 45°
    ('0.9D-1.4Y', 'Combination', 'Min', 'Top', 8.0, 8.0, -1),          # 45°, P tracción
    ('0.9D-1.4Y', 'Combination', 'Min', 'Bottom', 6.0, 6.0, -1),       # 45°, P tracción

    # Tracción con M2 dominante
    ('0.9D-1.4E_Y', 'Combination', 'Min', 'Top', 12.0, 4.0, -1),       # ~71.6°, P tracción
    ('0.9D-1.4E_Y', 'Combination', 'Min', 'Bottom', 10.0, 3.0, -1),    # ~73.3°, P tracción

    # Tracción alta con momento pequeño
    ('0.9D-1.6W', 'Combination', 'Min', 'Top', 3.0, 5.0, -1),          # ~31°, P tracción alta
    ('0.9D-1.6W', 'Combination', 'Min', 'Bottom', 2.5, 4.0, -1),       # ~32°, P tracción alta
]

for pier_name, (P_base, V2_base, V3_base, T_base) in pier_base.items():
    for combo, ctype, step, location, M2_factor, M3_factor, P_sign in combos_with_angles:
        # Escalar los factores según el pier
        scale = abs(P_base) / 40  # Normalizar usando P de PL1 como referencia
        M2 = M2_factor * scale * 0.3  # Factor de escala para momentos razonables
        M3 = M3_factor * scale * 0.3

        # Aplicar signo de P (positivo = tracción, negativo = compresión)
        P = abs(P_base) * P_sign * (-1)  # P_sign=-1 da tracción (P positivo)
        P_loc = P if location == 'Top' else P * 0.85

        # Variar signos de V y M según la combinación
        # Las combinaciones 'Min' y con signo negativo invierten cortantes
        V_sign = -1 if 'Min' in step or '-1.4' in combo else 1
        M_sign = -1 if '-1.4' in combo else 1

        V2_loc = V2_base * V_sign if location == 'Top' else V2_base * V_sign * 0.8
        V3_loc = V3_base * V_sign if location == 'Top' else V3_base * V_sign * 0.8
        T_loc = T_base if location == 'Top' else T_base * 0.7

        # Aplicar signo a momentos (mantener magnitud pero variar signo según combo)
        M2_final = M2 * M_sign
        M3_final = M3 * M_sign

        pier_forces.append(('Piso1', pier_name, combo, ctype, step, location,
                           P_loc, V2_loc, V3_loc, T_loc, M2_final, M3_final))

for row_idx, force in enumerate(pier_forces, 4):
    for col, val in enumerate(force, 1):
        ws5.cell(row=row_idx, column=col, value=val)

print(f'Pier Forces: {len(pier_forces)} filas')

# ============================================================================
# Sheet 6: Material Properties - Concrete
# ============================================================================
ws6 = wb.create_sheet('Material Properties - Concrete')
ws6['A1'] = 'TABLE:  Material Properties - Concrete'

headers6 = ['Name', 'Type', 'E', 'fc', 'fy']
for col, h in enumerate(headers6, 1):
    ws6.cell(row=2, column=col, value=h)

units6 = ['', '', 'tonf/m²', 'tonf/m²', 'tonf/m²']
for col, u in enumerate(units6, 1):
    ws6.cell(row=3, column=col, value=u)

# 3000 Psi = 21 MPa ≈ 2100 tonf/m²
materials = [
    ('3000Psi', 'Concrete', 2500000, 2100, 42000),
]
for row_idx, mat in enumerate(materials, 4):
    for col, val in enumerate(mat, 1):
        ws6.cell(row=row_idx, column=col, value=val)

# Guardar
output_path = 'examples/edge_cases_composite_piers.xlsx'
wb.save(output_path)
print(f'\nArchivo guardado: {output_path}')
print('\nPiers creados:')
print('  - PL1: L-shape (1200x200 + 800x200)')
print('  - PL2: L-shape invertida (1500x250 + 1000x250)')
print('  - PT1: T-shape (1400x200 + 1000x200)')
print('  - PT2: T-shape grande (2000x200 + 1200x250)')
print('  - PC1: C-shape (1500x200 + 2x800x200)')
print('  - PC2: C-shape grande (2000x250 + 2x1000x250)')
print('  - P1:  Rectangular (1000x200)')
