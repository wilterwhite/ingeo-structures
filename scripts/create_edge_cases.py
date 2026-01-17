"""
Script unificado para crear archivo de casos edge con todos los tipos de elementos.

Incluye:
- Piers compuestos (L, T, C) con geometría de muros
- Piers rectangulares con diferentes espesores (100-300mm)
- Columnas (diferentes tamaños incluyendo struts)
- Vigas
- Spandrels/Drop Beams
"""
import math
from openpyxl import Workbook

wb = Workbook()

# ============================================================================
# Helpers para geometría compuesta (FORMATO ETABS 3D VERTICAL)
# ============================================================================
# En ETABS, los muros son rectángulos 3D verticales:
# - Los 4 puntos definen un rectángulo donde:
#   - Pt1 y Pt2 están en Z_base (borde inferior del muro)
#   - Pt3 y Pt4 están en Z_top (borde superior del muro)
# - En planta XY, Pt1==Pt4 y Pt2==Pt3 (el muro es una LINEA en XY)
# - El espesor NO se ve en las coordenadas, solo en Area y Perimeter
# ============================================================================
all_points = {}
all_walls = []
all_pier_assigns = []
point_id_counter = 1
wall_id_counter = 1

# Niveles Z para el piso
Z_BASE = 0.0    # Base del muro (m)
Z_TOP = 3.0     # Tope del muro (m)
WALL_HEIGHT = Z_TOP - Z_BASE  # Altura del muro


def create_wall_points_3d(cx, cy, length, thickness, angle_deg=0):
    """
    Crea 4 puntos para un muro rectangular 3D VERTICAL (formato ETABS).

    En ETABS, un muro vertical tiene:
    - Pt1: extremo 1 de la linea central, Z=base
    - Pt2: extremo 2 de la linea central, Z=base
    - Pt3: extremo 2 de la linea central, Z=tope (mismo XY que Pt2)
    - Pt4: extremo 1 de la linea central, Z=tope (mismo XY que Pt1)

    El espesor se refleja en Area y Perimeter, NO en las coordenadas.

    Returns:
        Lista de 4 tuplas (x, y, z)
    """
    angle = math.radians(angle_deg)
    dx, dy = math.cos(angle), math.sin(angle)
    half_L = length / 2

    # Extremos de la linea central del muro en planta
    x1 = cx - dx * half_L
    y1 = cy - dy * half_L
    x2 = cx + dx * half_L
    y2 = cy + dy * half_L

    # 4 puntos del rectangulo 3D vertical
    return [
        (x1, y1, Z_BASE),  # Pt1: extremo 1, base
        (x2, y2, Z_BASE),  # Pt2: extremo 2, base
        (x2, y2, Z_TOP),   # Pt3: extremo 2, tope (mismo XY que Pt2)
        (x1, y1, Z_TOP),   # Pt4: extremo 1, tope (mismo XY que Pt1)
    ]


def add_point(x, y, z):
    """Agrega punto 3D y retorna ID."""
    global point_id_counter
    x, y, z = round(x, 4), round(y, 4), round(z, 4)

    for pid, (px, py, pz) in all_points.items():
        if abs(px - x) < 0.001 and abs(py - y) < 0.001 and abs(pz - z) < 0.001:
            return pid

    pid = point_id_counter
    all_points[pid] = (x, y, z)
    point_id_counter += 1
    return pid


def add_wall(story, pier_name, cx, cy, length, thickness, angle_deg=0):
    """
    Agrega muro rectangular 3D vertical y lo asigna a un pier.

    El muro se define por su linea central en planta (cx, cy, length, angle)
    y su espesor (thickness). La geometria 3D se genera con puntos en Z_BASE y Z_TOP.

    El Area y Perimeter se calculan como ETABS:
    - Area = length * WALL_HEIGHT (area del muro en m2)
    - Perimeter = 2 * (length + WALL_HEIGHT) (perimetro del muro en m)
    """
    global wall_id_counter

    points = create_wall_points_3d(cx, cy, length, thickness, angle_deg)
    pt_ids = [add_point(p[0], p[1], p[2]) for p in points]

    wall_id = wall_id_counter
    wall_bay = f"W{wall_id}"

    # Calcular Area y Perimeter como ETABS REALMENTE lo hace:
    # - Area = length * WALL_HEIGHT (superficie del muro en m2)
    # - Perimeter = 2 * (length + WALL_HEIGHT) (perimetro del rectangulo del muro en m)
    # NOTA: El espesor NO esta en Area/Perimeter, viene de Pier Section Properties
    area = length * WALL_HEIGHT  # Superficie del muro (m2)
    perimeter = 2 * (length + WALL_HEIGHT)  # Perimetro del rectangulo del muro (m)

    all_walls.append((wall_id, story, wall_bay, pt_ids[0], pt_ids[1], pt_ids[2], pt_ids[3], perimeter, area))
    all_pier_assigns.append((story, wall_bay, wall_id, pier_name))

    wall_id_counter += 1
    return wall_id


# ============================================================================
# Definir piers compuestos (L, T, C)
# ============================================================================
story = "Piso1"

# PL1: L-shape (1200x200 horizontal + 800x200 vertical)
add_wall(story, "PL1", cx=0.6, cy=0, length=1.2, thickness=0.2, angle_deg=0)
add_wall(story, "PL1", cx=1.2, cy=0.4, length=0.8, thickness=0.2, angle_deg=90)

# PL2: L-shape invertida (1500x250 + 1000x250)
add_wall(story, "PL2", cx=4.75, cy=0, length=1.5, thickness=0.25, angle_deg=0)
add_wall(story, "PL2", cx=4.0, cy=0.5, length=1.0, thickness=0.25, angle_deg=90)

# PT1: T-shape (1400x200 + 1000x200)
add_wall(story, "PT1", cx=8.0, cy=0, length=1.4, thickness=0.2, angle_deg=0)
add_wall(story, "PT1", cx=8.0, cy=-0.5, length=1.0, thickness=0.2, angle_deg=90)

# PT2: T-shape grande (2000x200 + 1200x250)
add_wall(story, "PT2", cx=12.0, cy=0, length=2.0, thickness=0.2, angle_deg=0)
add_wall(story, "PT2", cx=12.0, cy=-0.6, length=1.2, thickness=0.25, angle_deg=90)

# PC1: C-shape (1500x200 + 2x800x200)
add_wall(story, "PC1", cx=16.0, cy=0, length=1.5, thickness=0.2, angle_deg=90)
add_wall(story, "PC1", cx=16.4, cy=0.75, length=0.8, thickness=0.2, angle_deg=0)
add_wall(story, "PC1", cx=16.4, cy=-0.75, length=0.8, thickness=0.2, angle_deg=0)

# PC2: C-shape grande (2000x250 + 2x1000x250)
add_wall(story, "PC2", cx=20.0, cy=0, length=2.0, thickness=0.25, angle_deg=90)
add_wall(story, "PC2", cx=20.5, cy=1.0, length=1.0, thickness=0.25, angle_deg=0)
add_wall(story, "PC2", cx=20.5, cy=-1.0, length=1.0, thickness=0.25, angle_deg=0)

# P1: Rectangular simple (1000x200)
add_wall(story, "P1", cx=24.0, cy=0, length=1.0, thickness=0.2, angle_deg=0)

print(f"Geometría compuesta: {len(all_points)} puntos, {len(all_walls)} muros")

# ============================================================================
# Sheet 1: Point Object Connectivity
# ============================================================================
ws1 = wb.active
ws1.title = 'Point Object Connectivity'
ws1['A1'] = 'TABLE:  Point Object Connectivity'

headers1 = ['UniqueName', 'Is Auto Point', 'Story', 'PointBay', 'IsSpecial', 'X', 'Y', 'Z', 'GUID']
for col, h in enumerate(headers1, 1):
    ws1.cell(row=2, column=col, value=h)
units1 = ['', '', '', '', '', 'm', 'm', 'm', '']
for col, u in enumerate(units1, 1):
    ws1.cell(row=3, column=col, value=u)

for row_idx, (pid, (x, y, z)) in enumerate(sorted(all_points.items()), 4):
    ws1.cell(row=row_idx, column=1, value=pid)
    ws1.cell(row=row_idx, column=2, value='No')
    ws1.cell(row=row_idx, column=3, value='Piso1')
    ws1.cell(row=row_idx, column=4, value=pid)
    ws1.cell(row=row_idx, column=5, value='No')
    ws1.cell(row=row_idx, column=6, value=x)
    ws1.cell(row=row_idx, column=7, value=y)
    ws1.cell(row=row_idx, column=8, value=z)
    ws1.cell(row=row_idx, column=9, value=f'guid-{pid}')

# ============================================================================
# Sheet 2: Wall Object Connectivity
# ============================================================================
ws2 = wb.create_sheet('Wall Object Connectivity')
ws2['A1'] = 'TABLE:  Wall Object Connectivity'

headers2 = ['UniqueName', 'Story', 'WallBay', 'UniquePt1', 'UniquePt2', 'UniquePt3', 'UniquePt4', 'Perimeter', 'Area', 'GUID']
for col, h in enumerate(headers2, 1):
    ws2.cell(row=2, column=col, value=h)
units2 = ['', '', '', '', '', '', '', 'm', 'm2', '']
for col, u in enumerate(units2, 1):
    ws2.cell(row=3, column=col, value=u)

for row_idx, (wall_id, story, wall_bay, pt1, pt2, pt3, pt4, perimeter, area) in enumerate(all_walls, 4):
    ws2.cell(row=row_idx, column=1, value=wall_id)
    ws2.cell(row=row_idx, column=2, value=story)
    ws2.cell(row=row_idx, column=3, value=wall_bay)
    ws2.cell(row=row_idx, column=4, value=pt1)
    ws2.cell(row=row_idx, column=5, value=pt2)
    ws2.cell(row=row_idx, column=6, value=pt3)
    ws2.cell(row=row_idx, column=7, value=pt4)
    ws2.cell(row=row_idx, column=8, value=perimeter)
    ws2.cell(row=row_idx, column=9, value=area)
    ws2.cell(row=row_idx, column=10, value=f'wall-guid-{wall_id}')

# ============================================================================
# Sheet 3: Area Assigns - Pier Labels
# ============================================================================
ws3 = wb.create_sheet('Area Assigns - Pier Labels')
ws3['A1'] = 'TABLE:  Area Assignments - Pier Labels'

headers3 = ['Story', 'Label', 'UniqueName', 'Pier Name']
for col, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=col, value=h)

for row_idx, (story, wall_label, wall_id, pier_name) in enumerate(all_pier_assigns, 3):
    ws3.cell(row=row_idx, column=1, value=story)
    ws3.cell(row=row_idx, column=2, value=wall_label)
    ws3.cell(row=row_idx, column=3, value=wall_id)
    ws3.cell(row=row_idx, column=4, value=pier_name)

# ============================================================================
# Sheet 4: Pier Section Properties (compuestos + diferentes espesores)
# Formato idéntico a ETABS (con CG Bottom X/Y/Z, CG Top X/Y/Z)
# ============================================================================
ws4 = wb.create_sheet('Pier Section Properties')
ws4['A1'] = 'TABLE:  Pier Section Properties'

headers4 = ['Story', 'Pier', 'AxisAngle', '# Area Objects', '# Line Objects',
            'Width Bottom', 'Thickness Bottom', 'Width Top', 'Thickness Top', 'Material',
            'CG Bottom X', 'CG Bottom Y', 'CG Bottom Z', 'CG Top X', 'CG Top Y', 'CG Top Z']
for col, h in enumerate(headers4, 1):
    ws4.cell(row=2, column=col, value=h)
units4 = ['', '', 'deg', '', '', 'm', 'm', 'm', 'm', '', 'm', 'm', 'm', 'm', 'm', 'm']
for col, u in enumerate(units4, 1):
    ws4.cell(row=3, column=col, value=u)

# (story, pier, angle, n_area, n_line, w_bot, t_bot, w_top, t_top, mat, cg_bot_x, cg_bot_y, cg_bot_z, cg_top_x, cg_top_y, cg_top_z)
pier_props = [
    # Compuestos
    ('Piso1', 'PL1', 0, 2, 0, 1.2, 0.20, 1.2, 0.20, '3000Psi', 0.7, 0.2, 0, 0.7, 0.2, 3.0),
    ('Piso1', 'PL2', 0, 2, 0, 1.5, 0.25, 1.5, 0.25, '3000Psi', 4.5, 0.25, 0, 4.5, 0.25, 3.0),
    ('Piso1', 'PT1', 0, 2, 0, 1.4, 0.20, 1.4, 0.20, '3000Psi', 8.0, -0.25, 0, 8.0, -0.25, 3.0),
    ('Piso1', 'PT2', 0, 2, 0, 2.0, 0.25, 2.0, 0.25, '3000Psi', 12.0, -0.3, 0, 12.0, -0.3, 3.0),
    ('Piso1', 'PC1', 0, 3, 0, 1.5, 0.20, 1.5, 0.20, '3000Psi', 16.2, 0.0, 0, 16.2, 0.0, 3.0),
    ('Piso1', 'PC2', 0, 3, 0, 2.0, 0.25, 2.0, 0.25, '3000Psi', 20.25, 0.0, 0, 20.25, 0.0, 3.0),
    ('Piso1', 'P1', 0, 1, 0, 1.0, 0.20, 1.0, 0.20, '3000Psi', 24.0, 0.0, 0, 24.0, 0.0, 3.0),
    # Diferentes espesores (para probar enfierradura mínima)
    ('Piso1', 'P_e10', 0, 1, 0, 0.8, 0.10, 0.8, 0.10, '3000Psi', 28.0, 0.0, 0, 28.0, 0.0, 3.0),  # e=100mm (malla simple)
    ('Piso1', 'P_e12', 0, 1, 0, 0.9, 0.12, 0.9, 0.12, '3000Psi', 30.0, 0.0, 0, 30.0, 0.0, 3.0),  # e=120mm (malla simple)
    ('Piso1', 'P_e13', 0, 1, 0, 1.0, 0.13, 1.0, 0.13, '3000Psi', 32.0, 0.0, 0, 32.0, 0.0, 3.0),  # e=130mm (límite malla simple)
    ('Piso1', 'P_e15', 0, 1, 0, 1.2, 0.15, 1.2, 0.15, '3000Psi', 34.0, 0.0, 0, 34.0, 0.0, 3.0),  # e=150mm (malla doble)
    ('Piso1', 'P_e20', 0, 1, 0, 1.5, 0.20, 1.5, 0.20, '3000Psi', 36.0, 0.0, 0, 36.0, 0.0, 3.0),  # e=200mm (malla doble)
    ('Piso1', 'P_e25', 0, 1, 0, 2.0, 0.25, 2.0, 0.25, '3000Psi', 38.0, 0.0, 0, 38.0, 0.0, 3.0),  # e=250mm (malla doble)
    ('Piso1', 'P_e30', 0, 1, 0, 2.5, 0.30, 2.5, 0.30, '3000Psi', 40.0, 0.0, 0, 40.0, 0.0, 3.0),  # e=300mm (malla doble)
]
for row_idx, pier in enumerate(pier_props, 4):
    for col, val in enumerate(pier, 1):
        ws4.cell(row=row_idx, column=col, value=val)

# ============================================================================
# Sheet 5: Pier Forces
# ============================================================================
ws5 = wb.create_sheet('Pier Forces')
ws5['A1'] = 'TABLE:  Pier Forces'

headers5 = ['Story', 'Pier', 'Output Case', 'Case Type', 'Step Type', 'Location',
            'P', 'V2', 'V3', 'T', 'M2', 'M3']
for col, h in enumerate(headers5, 1):
    ws5.cell(row=2, column=col, value=h)
units5 = ['', '', '', '', '', '', 'tonf', 'tonf', 'tonf', 'tonf-m', 'tonf-m', 'tonf-m']
for col, u in enumerate(units5, 1):
    ws5.cell(row=3, column=col, value=u)

# Fuerzas base [P_top, V2_top, M3_top, P_bot, V2_bot, M3_bot]
pier_forces_base = {
    'PL1': (-40, 10, 19, -34, 8, 15),
    'PL2': (-60, 15, 40, -52, 12, 32),
    'PT1': (-50, 12, 110, -42, 10, 88),
    'PT2': (-80, 20, 65, -70, 16, 52),
    'PC1': (-70, 18, 147, -60, 14, 118),
    'PC2': (-100, 25, 107, -88, 20, 86),
    'P1': (-35, 8, 43, -30, 6, 34),
    'P_e10': (-15, 5, 10, -12, 4, 8),
    'P_e12': (-20, 6, 12, -16, 5, 10),
    'P_e13': (-25, 7, 15, -20, 6, 12),
    'P_e15': (-35, 10, 25, -28, 8, 20),
    'P_e20': (-50, 14, 40, -42, 11, 32),
    'P_e25': (-75, 20, 70, -65, 16, 56),
    'P_e30': (-100, 28, 100, -85, 22, 80),
}

combos = [
    ('1.2D+1.6L', 'Combination', ''),
    ('1.4X+1.2D+1.0L', 'Combination', 'Max'),
    ('1.4Y+1.2D+1.0L', 'Combination', 'Max'),
    ('1.2D+1.0E_30', 'Combination', 'Max'),
    ('1.2D+1.0E_60', 'Combination', 'Max'),
    ('1.2D+1.0E_90', 'Combination', 'Max'),
]

pier_forces = []
for pier_name, (P_top, V2_top, M3_top, P_bot, V2_bot, M3_bot) in pier_forces_base.items():
    for combo, ctype, step in combos:
        pier_forces.append(('Piso1', pier_name, combo, ctype, step, 'Top', P_top, V2_top, 2, 0.5, 0.3, M3_top))
        pier_forces.append(('Piso1', pier_name, combo, ctype, step, 'Bottom', P_bot, V2_bot, 1.6, 0.35, 0.15, M3_bot))

for row_idx, force in enumerate(pier_forces, 4):
    for col, val in enumerate(force, 1):
        ws5.cell(row=row_idx, column=col, value=val)

print(f'Pier Forces: {len(pier_forces)} filas')

# ============================================================================
# Sheet 6: Wall Property Def - Specified (requerido por parser)
# ============================================================================
ws6 = wb.create_sheet('Wall Property Def - Specified')
ws6['A1'] = 'TABLE:  Wall Property Definitions - Specified'

headers6 = ['Name', 'Material', 'Type', 'Thickness']
for col, h in enumerate(headers6, 1):
    ws6.cell(row=2, column=col, value=h)
units6 = ['', '', '', 'm']
for col, u in enumerate(units6, 1):
    ws6.cell(row=3, column=col, value=u)

wall_props = [
    ('Wall20', '3000Psi', 'Specified', 0.20),
    ('Wall25', '3000Psi', 'Specified', 0.25),
    ('Wall26', '3000Psi', 'Specified', 0.26),
]
for row_idx, wall in enumerate(wall_props, 4):
    for col, val in enumerate(wall, 1):
        ws6.cell(row=row_idx, column=col, value=val)

# ============================================================================
# Sheet 7: Spandrel Section Properties (Drop Beams)
# Formato idéntico a ETABS (con CG Left/Right X/Y/Z)
# ============================================================================
ws7 = wb.create_sheet('Spandrel Section Properties')
ws7['A1'] = 'TABLE:  Spandrel Section Properties'

headers7 = ['Story', 'Spandrel', '# Area Objects', '# Line Objects', 'Length',
            'Depth Left', 'Thickness Left', 'Depth Right', 'Thickness Right', 'Material',
            'CG Left X', 'CG Left Y', 'CG Left Z', 'CG Right X', 'CG Right Y', 'CG Right Z']
for col, h in enumerate(headers7, 1):
    ws7.cell(row=2, column=col, value=h)
units7 = ['', '', '', '', 'm', 'm', 'm', 'm', 'm', '', 'm', 'm', 'm', 'm', 'm', 'm']
for col, u in enumerate(units7, 1):
    ws7.cell(row=3, column=col, value=u)

# (story, spandrel, n_area, n_line, length, depth_l, thick_l, depth_r, thick_r, mat, cg_l_x, cg_l_y, cg_l_z, cg_r_x, cg_r_y, cg_r_z)
spandrels = [
    ('Piso1', 'S1', 2, 0, 0.45, 0.66, 0.26, 0.66, 0.26, '3000Psi', 11.4, 2.85, 2.5, 11.4, 3.3, 2.5),
    ('Piso1', 'S2', 4, 0, 0.90, 0.66, 0.21, 0.66, 0.21, '3000Psi', 9.6, 2.85, 2.5, 9.6, 3.75, 2.5),
    ('Piso1', 'S3', 3, 0, 1.20, 0.40, 0.20, 0.40, 0.20, '3000Psi', 5.0, 2.0, 2.5, 5.0, 3.2, 2.5),
]
for row_idx, span in enumerate(spandrels, 4):
    for col, val in enumerate(span, 1):
        ws7.cell(row=row_idx, column=col, value=val)

# ============================================================================
# Sheet 8: Spandrel Forces
# ============================================================================
ws8 = wb.create_sheet('Spandrel Forces')
ws8['A1'] = 'TABLE:  Spandrel Forces'

headers8 = ['Story', 'Spandrel', 'Output Case', 'Case Type', 'Step Type', 'Location',
            'P', 'V2', 'V3', 'T', 'M2', 'M3']
for col, h in enumerate(headers8, 1):
    ws8.cell(row=2, column=col, value=h)
units8 = ['', '', '', '', '', '', 'tonf', 'tonf', 'tonf', 'tonf-m', 'tonf-m', 'tonf-m']
for col, u in enumerate(units8, 1):
    ws8.cell(row=3, column=col, value=u)

spandrel_forces = []
combos_simple = [('1.2D+1.6L', 'Combination', ''), ('1.4X+1.2D+1.0L', 'Combination', 'Max'), ('1.4Y+1.2D+1.0L', 'Combination', 'Max')]

for combo, ctype, step in combos_simple:
    spandrel_forces.append(('Piso1', 'S1', combo, ctype, step, 'Left', -6, 8, 0.2, 0.1, 0.1, 2))
    spandrel_forces.append(('Piso1', 'S1', combo, ctype, step, 'Right', -6, 8, 0.2, 0.1, 0.1, 3))
    spandrel_forces.append(('Piso1', 'S2', combo, ctype, step, 'Left', -8, 10, 0.3, 0.15, 0.15, 3))
    spandrel_forces.append(('Piso1', 'S2', combo, ctype, step, 'Right', -8, 10, 0.3, 0.15, 0.15, 4))
    spandrel_forces.append(('Piso1', 'S3', combo, ctype, step, 'Left', -5, 6, 0.15, 0.1, 0.1, 2))
    spandrel_forces.append(('Piso1', 'S3', combo, ctype, step, 'Right', -5, 6, 0.15, 0.1, 0.1, 2.5))

for row_idx, force in enumerate(spandrel_forces, 4):
    for col, val in enumerate(force, 1):
        ws8.cell(row=row_idx, column=col, value=val)

print(f'Spandrel Forces: {len(spandrel_forces)} filas')

# ============================================================================
# Sheet 9: Frame Sec Def - Conc Rect (Columns + Beams)
# ============================================================================
ws9 = wb.create_sheet('Frame Sec Def - Conc Rect')
ws9['A1'] = 'TABLE:  Frame Section Property Definitions - Concrete Rectangular'

headers9 = ['Name', 'Material', 'From File?', 'Depth', 'Width', 'Design Type']
for col, h in enumerate(headers9, 1):
    ws9.cell(row=2, column=col, value=h)
units9 = ['', '', '', 'm', 'm', '']
for col, u in enumerate(units9, 1):
    ws9.cell(row=3, column=col, value=u)

frame_sections = [
    # Struts cuadrados (ambas dim < 15cm)
    ('COL_12x12', 'Concreto28', 'No', 0.12, 0.12, 'Column'),
    ('COL_14x14', 'Concreto28', 'No', 0.14, 0.14, 'Column'),
    # Struts rectangulares (dim menor < 15cm, ratio < 3)
    ('COL_14x19', 'Concreto28', 'No', 0.19, 0.14, 'Column'),  # Strut: 140mm < 150mm, ratio=1.36 < 3
    ('COL_19x14', 'Concreto28', 'No', 0.14, 0.19, 'Column'),  # Strut: 140mm < 150mm, ratio=1.36 < 3
    ('COL_14x28', 'Concreto28', 'No', 0.28, 0.14, 'Column'),  # Strut: 140mm < 150mm, ratio=2.0 < 3
    # Limite strut/muro (dim menor < 15cm, ratio >= 3 -> muro)
    ('COL_14x42', 'Concreto28', 'No', 0.42, 0.14, 'Column'),  # Muro: 140mm < 150mm, ratio=3.0 >= 3
    ('COL_15x12', 'Concreto28', 'No', 0.15, 0.12, 'Column'),
    # Columnas normales (dim menor >= 15cm)
    ('COL_15x15', 'Concreto28', 'No', 0.15, 0.15, 'Column'),
    ('COL_20x20', 'Concreto28', 'No', 0.20, 0.20, 'Column'),
    ('COL_40x20', 'Concreto28', 'No', 0.40, 0.20, 'Column'),
    # Vigas
    ('V12x29', 'Concreto28', 'No', 0.29, 0.12, 'Beam'),
    ('V20x35', 'Concreto28', 'No', 0.35, 0.20, 'Beam'),
    ('V20x70', 'Concreto28', 'No', 0.70, 0.20, 'Beam'),
]
for row_idx, sec in enumerate(frame_sections, 4):
    for col, val in enumerate(sec, 1):
        ws9.cell(row=row_idx, column=col, value=val)

# ============================================================================
# Sheet 10: Frame Assigns - Sect Prop
# ============================================================================
ws10 = wb.create_sheet('Frame Assigns - Sect Prop')
ws10['A1'] = 'TABLE:  Frame Assignments - Section Properties'

headers10 = ['Story', 'Label', 'Section Property']
for col, h in enumerate(headers10, 1):
    ws10.cell(row=2, column=col, value=h)

assigns = [
    # Struts cuadrados
    ('Piso1', 'C1', 'COL_12x12'),
    ('Piso1', 'C2', 'COL_14x14'),
    # Struts rectangulares (nuevos casos)
    ('Piso1', 'C3', 'COL_14x19'),   # Strut rectangular
    ('Piso1', 'C4', 'COL_19x14'),   # Strut rectangular
    ('Piso1', 'C5', 'COL_14x28'),   # Strut rectangular (ratio=2.0)
    ('Piso1', 'C6', 'COL_14x42'),   # Muro delgado (ratio=3.0)
    ('Piso1', 'C7', 'COL_15x12'),   # Strut (no porque 150mm >= limite)
    # Columnas normales
    ('Piso1', 'C8', 'COL_15x15'),
    ('Piso1', 'C9', 'COL_20x20'),
    ('Piso1', 'C10', 'COL_40x20'),
    # Vigas
    ('Piso1', 'B1', 'V12x29'),
    ('Piso1', 'B2', 'V20x35'),
    ('Piso1', 'B3', 'V20x70'),
]
for row_idx, assign in enumerate(assigns, 3):
    for col, val in enumerate(assign, 1):
        ws10.cell(row=row_idx, column=col, value=val)

# ============================================================================
# Sheet 11: Element Forces - Columns
# ============================================================================
ws11 = wb.create_sheet('Element Forces - Columns')
ws11['A1'] = 'TABLE:  Element Forces - Columns'

headers11 = ['Story', 'Column', 'Unique Name', 'Output Case', 'Case Type', 'Step Type',
             'Station', 'P', 'V2', 'V3', 'T', 'M2', 'M3']
for col, h in enumerate(headers11, 1):
    ws11.cell(row=2, column=col, value=h)
units11 = ['', '', '', '', '', '', 'm', 'tonf', 'tonf', 'tonf', 'tonf-m', 'tonf-m', 'tonf-m']
for col, u in enumerate(units11, 1):
    ws11.cell(row=3, column=col, value=u)

col_forces = []
height = 2.8
col_loads = {
    # Struts cuadrados
    'C1': [(-5, 0.5, 0.3, 0.5, 0.8), (-3, 0.8, 0.5, 0.6, 1.0), (-4, 0.6, 0.4, 0.55, 0.9)],
    'C2': [(-8, 0.6, 0.4, 0.6, 1.0), (-6, 1.0, 0.6, 0.8, 1.2), (-7, 0.8, 0.5, 0.7, 1.1)],
    # Struts rectangulares
    'C3': [(-10, 0.8, 0.5, 0.7, 1.1), (-8, 1.2, 0.7, 0.9, 1.4), (-9, 1.0, 0.6, 0.8, 1.2)],
    'C4': [(-12, 0.9, 0.6, 0.8, 1.3), (-10, 1.4, 0.8, 1.0, 1.6), (-11, 1.2, 0.7, 0.9, 1.4)],
    'C5': [(-15, 1.2, 0.8, 1.0, 1.8), (-12, 1.8, 1.0, 1.2, 2.2), (-14, 1.5, 0.9, 1.1, 2.0)],
    # Muro delgado (ratio=3)
    'C6': [(-25, 2.0, 1.2, 1.5, 3.0), (-20, 3.0, 1.5, 1.8, 4.0), (-22, 2.5, 1.4, 1.6, 3.5)],
    'C7': [(-20, 2.0, 1.0, 1.2, 2.0), (-15, 3.0, 1.5, 1.5, 2.5), (-18, 2.5, 1.2, 1.3, 2.2)],
    # Columnas normales
    'C8': [(-30, 2.5, 1.5, 1.5, 2.5), (-25, 4.0, 2.0, 2.0, 3.5), (-28, 3.0, 1.8, 1.8, 3.0)],
    'C9': [(-50, 3.0, 2.0, 2.0, 3.5), (-40, 5.0, 3.0, 2.5, 5.0), (-45, 4.0, 2.5, 2.2, 4.2)],
    'C10': [(-100, 5.0, 3.0, 3.0, 6.0), (-80, 8.0, 5.0, 4.0, 8.0), (-90, 6.5, 4.0, 3.5, 7.0)],
}

unique_id = 1000
for col_name, loads in col_loads.items():
    for i, (combo, ctype, step) in enumerate(combos_simple):
        P, V2, V3, M2, M3 = loads[i]
        col_forces.append(('Piso1', col_name, str(unique_id), combo, ctype, step, 0, P, V2, V3, 0.1, M2, M3))
        col_forces.append(('Piso1', col_name, str(unique_id), combo, ctype, step, height, P*0.9, V2*0.9, V3*0.9, 0.08, M2*0.8, M3*0.8))
    unique_id += 1

for row_idx, force in enumerate(col_forces, 4):
    for col, val in enumerate(force, 1):
        ws11.cell(row=row_idx, column=col, value=val)

print(f'Column Forces: {len(col_forces)} filas')

# ============================================================================
# Sheet 12: Element Forces - Beams
# ============================================================================
ws12 = wb.create_sheet('Element Forces - Beams')
ws12['A1'] = 'TABLE:  Element Forces - Beams'

headers12 = ['Story', 'Beam', 'Unique Name', 'Output Case', 'Case Type', 'Step Type',
             'Station', 'P', 'V2', 'V3', 'T', 'M2', 'M3']
for col, h in enumerate(headers12, 1):
    ws12.cell(row=2, column=col, value=h)
units12 = ['', '', '', '', '', '', 'm', 'tonf', 'tonf', 'tonf', 'tonf-m', 'tonf-m', 'tonf-m']
for col, u in enumerate(units12, 1):
    ws12.cell(row=3, column=col, value=u)

beam_forces = []
beam_loads = {
    'B1': [(0, 3.5, 4.0), (0, 4.0, 5.0), (0, 3.8, 4.5)],
    'B2': [(0, 5.0, 6.0), (0, 6.0, 8.0), (0, 5.5, 7.0)],
    'B3': [(0, 8.0, 10.0), (0, 10.0, 12.0), (0, 9.0, 11.0)],
}
beam_lengths = {'B1': 3.0, 'B2': 4.0, 'B3': 6.0}

unique_id = 2000
for beam_name, loads in beam_loads.items():
    length = beam_lengths[beam_name]
    for i, (combo, ctype, step) in enumerate(combos_simple):
        P, V2, M3 = loads[i]
        beam_forces.append(('Piso1', beam_name, str(unique_id), combo, ctype, step, 0, P, V2, 0, 0, 0, M3))
        beam_forces.append(('Piso1', beam_name, str(unique_id), combo, ctype, step, length/2, P, 0, 0, 0, 0, M3*1.3))
        beam_forces.append(('Piso1', beam_name, str(unique_id), combo, ctype, step, length, P, -V2, 0, 0, 0, M3))
    unique_id += 1

for row_idx, force in enumerate(beam_forces, 4):
    for col, val in enumerate(force, 1):
        ws12.cell(row=row_idx, column=col, value=val)

print(f'Beam Forces: {len(beam_forces)} filas')

# ============================================================================
# Sheet 13: Section Cut Forces - Analysis (Vigas Capitel / Drop Beams)
# ============================================================================
ws13 = wb.create_sheet('Section Cut Forces - Analysis')
ws13['A1'] = 'TABLE:  Section Cut Forces - Analysis'

headers13 = ['SectionCut', 'Output Case', 'Case Type', 'Step Type',
             'F1', 'F2', 'F3', 'M1', 'M2', 'M3', 'X', 'Y', 'Z']
for col, h in enumerate(headers13, 1):
    ws13.cell(row=2, column=col, value=h)
units13 = ['', '', '', '', 'tonf', 'tonf', 'tonf', 'tonf-m', 'tonf-m', 'tonf-m', 'm', 'm', 'm']
for col, u in enumerate(units13, 1):
    ws13.cell(row=3, column=col, value=u)

# Vigas capitel con diferentes dimensiones (extraídas del nombre)
# Formato: SCut-Losa STORY-WxH DESCRIPCION
section_cuts = [
    # Viga capitel 29x100 (ancho x alto)
    ('SCut-Losa Piso1-29x100 Borde losa-1', 5.0, 7.2, 3.0),
    ('SCut-Losa Piso1-29x100 Borde losa-2', 6.2, 9.4, 3.0),
    # Viga capitel 20x80
    ('SCut-Losa Piso1-20x80 Central-1', 10.0, 5.0, 3.0),
    # Viga capitel 25x60
    ('SCut-Losa Piso1-25x60 Esquina-1', 2.0, 2.0, 3.0),
]

section_cut_forces = []
combos_scut = [
    ('1.2D+1.6L', 'Combination', ''),
    ('1.4X+1.2D+1.0L', 'Combination', 'Max'),
    ('1.4Y+1.2D+1.0L', 'Combination', 'Max'),
]

for scut_name, x, y, z in section_cuts:
    for combo, ctype, step in combos_scut:
        # F1=axial, F2=corte menor, F3=corte mayor, M1=torsion, M2=momento menor, M3=momento mayor
        section_cut_forces.append((scut_name, combo, ctype, step, -0.5, 1.5, 8.0, 0.5, 2.0, 12.0, x, y, z))

for row_idx, force in enumerate(section_cut_forces, 4):
    for col, val in enumerate(force, 1):
        ws13.cell(row=row_idx, column=col, value=val)

print(f'Section Cut Forces: {len(section_cut_forces)} filas')

# ============================================================================
# Sheet 14: Material Properties - Concrete
# ============================================================================
ws14 = wb.create_sheet('Material Properties - Concrete')
ws14['A1'] = 'TABLE:  Material Properties - Concrete'

headers14 = ['Name', 'Type', 'E', 'fc', 'fy']
for col, h in enumerate(headers14, 1):
    ws14.cell(row=2, column=col, value=h)
units14 = ['', '', 'tonf/m2', 'tonf/m2', 'tonf/m2']
for col, u in enumerate(units14, 1):
    ws14.cell(row=3, column=col, value=u)

materials = [
    ('3000Psi', 'Concrete', 2500000, 2100, 42000),
    ('Concreto28', 'Concrete', 2500000, 2800, 42000),
]
for row_idx, mat in enumerate(materials, 4):
    for col, val in enumerate(mat, 1):
        ws14.cell(row=row_idx, column=col, value=val)

# ============================================================================
# Guardar
# ============================================================================
output_path = 'examples/edge_cases.xlsx'
wb.save(output_path)

print(f'\n{"="*60}')
print(f'Archivo guardado: {output_path}')
print(f'{"="*60}')
print('\nPiers compuestos (L, T, C):')
print('  - PL1, PL2: Seccion L')
print('  - PT1, PT2: Seccion T')
print('  - PC1, PC2: Seccion C')
print('  - P1: Rectangular simple')
print('\nPiers con diferentes espesores:')
print('  - P_e10: e=100mm (malla simple, 1M)')
print('  - P_e12: e=120mm (malla simple, 1M)')
print('  - P_e13: e=130mm (limite malla simple, 1M)')
print('  - P_e15: e=150mm (malla doble, 2M)')
print('  - P_e20: e=200mm (malla doble, 2M)')
print('  - P_e25: e=250mm (malla doble, 2M)')
print('  - P_e30: e=300mm (malla doble, 2M)')
print('\nColumnas:')
print('  - C1, C2: Struts cuadrados (12x12, 14x14)')
print('  - C3, C4, C5: Struts rectangulares (14x19, 19x14, 14x28) - ratio < 3')
print('  - C6: Muro delgado (14x42) - ratio = 3, NO es strut')
print('  - C7: Limite strut (15x12) - 150mm no es strut')
print('  - C8-C10: Columnas normales (15x15, 20x20, 40x20)')
print('\nVigas:')
print('  - B1, B2, B3: Diferentes tamaños')
print('\nSpandrels:')
print('  - S1, S2, S3: Diferentes espesores')
print('\nVigas Capitel (Section Cuts):')
print('  - SCut 29x100: Borde losa')
print('  - SCut 20x80: Central')
print('  - SCut 25x60: Esquina')
