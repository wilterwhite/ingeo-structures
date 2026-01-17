# app/services/parsing/table_extractor.py
"""
Extraccion de tablas desde archivos Excel de ETABS.

ETABS exporta multiples tablas en una sola hoja de Excel, cada una
marcada con "TABLE: NombreTabla". Este modulo contiene la logica
para encontrar y extraer tablas especificas.

Las tablas de ETABS tienen esta estructura:
- Fila 0: "TABLE: NombreTabla"
- Fila 1: Headers (nombres de columnas)
- Fila 2: Unidades (mm, m, tonf, etc.)
- Fila 3+: Datos
"""
from typing import Optional, Dict, List, Tuple
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook


# =============================================================================
# Factores de Conversion de Unidades a mm
# =============================================================================

UNIT_FACTORS_TO_MM: Dict[str, float] = {
    'mm': 1.0,
    'm': 1000.0,
    'cm': 10.0,
    'in': 25.4,
    'ft': 304.8,
}


def get_unit_factor(unit_str) -> float:
    """
    Retorna el factor para convertir una unidad de longitud a mm.

    Args:
        unit_str: String con la unidad (ej: 'mm', 'm', 'cm')

    Returns:
        Factor de conversion a mm. Si la unidad no es reconocida, retorna 1.0
    """
    if unit_str is None or pd.isna(unit_str):
        return 1.0
    unit = str(unit_str).lower().strip()
    return UNIT_FACTORS_TO_MM.get(unit, 1.0)


def extract_units_from_df(df: pd.DataFrame) -> Dict[str, float]:
    """
    Extrae los factores de conversion de unidades de un DataFrame de ETABS.

    ETABS pone las unidades en la primera fila de datos (despues de headers).
    Esta funcion lee esa fila y retorna un dict con los factores de conversion.

    Args:
        df: DataFrame ya normalizado con headers

    Returns:
        Dict[column_name -> factor_to_mm]
        Ejemplo: {'width bottom': 1.0, 'cg bottom z': 1000.0}
    """
    units = {}
    if df is None or len(df) == 0:
        return units

    # La primera fila del DataFrame normalizado deberia tener las unidades
    first_row = df.iloc[0]
    for col in df.columns:
        unit_value = first_row.get(col)
        units[str(col).lower()] = get_unit_factor(unit_value)

    return units


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza los nombres de columnas a minusculas.

    Args:
        df: DataFrame con columnas a normalizar

    Returns:
        DataFrame con columnas en minusculas
    """
    df = df.copy()
    df.columns = [str(col).lower().strip() for col in df.columns]
    return df


# =============================================================================
# Mapeo de Nombres de Tabla ETABS -> Claves Internas
# =============================================================================

# Mapeo de patrones de nombre de tabla ETABS a claves internas
# El orden importa: patrones mas especificos primero
TABLE_NAME_MAPPINGS: List[Tuple[str, str]] = [
    # Wall / Pier properties
    ('wall property def', 'wall_props'),
    ('pier section properties', 'pier_props'),
    ('pier forces', 'pier_forces'),

    # Frame sections (rectangular y circular)
    ('frame section property definitions - concrete rectangular', 'frame_section'),
    ('frame sec def - conc rect', 'frame_section'),
    ('frame section property definitions - concrete circle', 'frame_circular'),

    # Frame assignments
    ('frame assignments - section properties', 'frame_assigns'),
    ('frame assigns - sect prop', 'frame_assigns'),

    # Element forces
    ('element forces - columns', 'column_forces'),
    ('element forces - beams', 'beam_forces'),

    # Spandrels
    ('spandrel section properties', 'spandrel_props'),
    ('spandrel forces', 'spandrel_forces'),

    # Section cuts (vigas capitel)
    ('section cut forces - analysis', 'section_cut'),

    # Connectivity (para geometria compuesta L, T, C)
    ('wall object connectivity', 'walls_connectivity'),
    ('point object connectivity', 'points_connectivity'),

    # Pier assignments
    ('area assign', 'pier_assigns'),
    ('area assignments - pier labels', 'pier_assigns'),
]


def _map_table_name_to_key(table_name: str) -> Optional[str]:
    """
    Mapea el nombre completo de una tabla ETABS a su clave interna.

    Args:
        table_name: Nombre de la tabla como aparece en el Excel

    Returns:
        Clave interna (ej: 'pier_props') o None si no es una tabla conocida
    """
    name_lower = table_name.lower()

    for pattern, key in TABLE_NAME_MAPPINGS:
        if pattern in name_lower:
            return key

    return None


# =============================================================================
# Extraccion Rapida con openpyxl (read_only mode) + Paralelizacion AGRESIVA
# =============================================================================

import time
import logging
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

_perf_logger = logging.getLogger('perf')

# Usar todos los cores disponibles
MAX_WORKERS = os.cpu_count() or 8


def _process_sheet(file_content: bytes, sheet_name: str) -> Dict[str, pd.DataFrame]:
    """
    Procesa una hoja individual en su propio thread.

    Cada thread abre su propia copia del workbook para evitar conflictos.
    """
    tables: Dict[str, pd.DataFrame] = {}

    # Abrir workbook en modo read_only (cada thread su propia copia)
    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    ws = wb[sheet_name]

    t_sheet = time.perf_counter()

    # UNA SOLA PASADA: Cargar todas las filas y encontrar tablas simultáneamente
    all_rows: List[tuple] = []
    table_positions: List[Tuple[str, int]] = []  # (table_name, row_number)

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        all_rows.append(row)
        # Buscar "TABLE:" en la primera celda
        first_cell = row[0] if row else None
        if first_cell and isinstance(first_cell, str) and 'TABLE:' in first_cell.upper():
            match = re.search(r'TABLE:\s*(.+)', first_cell, re.IGNORECASE)
            if match:
                table_name = match.group(1).strip()
                table_positions.append((table_name, row_idx))

    t_read = time.perf_counter()
    _perf_logger.info(f"[PERF] Sheet '{sheet_name}': read {len(all_rows)} rows in {t_read-t_sheet:.2f}s")

    wb.close()

    if not table_positions:
        return tables

    # Extraer cada tabla
    for i, (table_name, start_row) in enumerate(table_positions):
        # Determinar fin de tabla
        if i + 1 < len(table_positions):
            end_row = table_positions[i + 1][1]
        else:
            end_row = len(all_rows)

        # start_row apunta a "TABLE:", start_row+1 son headers
        header_idx = start_row + 1
        if header_idx >= len(all_rows):
            continue

        headers = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1:end_row]

        if not data_rows:
            continue

        # Crear DataFrame
        df = pd.DataFrame(data_rows, columns=headers)

        # Eliminar filas completamente vacias
        df = df.dropna(how='all')

        if len(df) > 0:
            # Mapear nombre de tabla a clave interna
            table_key = _map_table_name_to_key(table_name)
            if table_key:
                tables[table_key] = df
                _perf_logger.info(f"[PERF] Table '{table_key}': {len(df)} rows")

    return tables


def extract_tables_fast(file_content: bytes) -> Dict[str, pd.DataFrame]:
    """
    Extrae todas las tablas de un archivo Excel usando openpyxl read_only mode.

    OPTIMIZADO AGRESIVO:
    - Todas las hojas en paralelo con ThreadPoolExecutor
    - Usa todos los cores disponibles (24 en este sistema)
    - ThreadPool evita overhead de serializacion de ProcessPool en Windows

    Args:
        file_content: Contenido binario del archivo Excel

    Returns:
        Dict[table_key -> DataFrame] con todas las tablas encontradas
    """
    t0 = time.perf_counter()
    tables: Dict[str, pd.DataFrame] = {}

    # Obtener lista de hojas (lectura rapida)
    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    t1 = time.perf_counter()
    _perf_logger.info(f"[PERF] get_sheetnames: {t1-t0:.2f}s ({len(sheet_names)} sheets)")

    # AGRESIVO: Procesar TODAS las hojas en paralelo con todos los workers
    n_workers = min(MAX_WORKERS, len(sheet_names))
    _perf_logger.info(f"[PERF] Procesando {len(sheet_names)} hojas en paralelo con {n_workers} workers (max={MAX_WORKERS})")

    with ThreadPoolExecutor(max_workers=n_workers) as executor:
        # Lanzar todas las hojas en paralelo
        future_to_sheet = {
            executor.submit(_process_sheet, file_content, sheet_name): sheet_name
            for sheet_name in sheet_names
        }

        # Recolectar resultados a medida que terminan
        for future in as_completed(future_to_sheet):
            sheet_name = future_to_sheet[future]
            try:
                sheet_tables = future.result()
                for key, df in sheet_tables.items():
                    if key not in tables:
                        tables[key] = df
            except Exception as e:
                _perf_logger.warning(f"[PERF] Error procesando {sheet_name}: {e}")

    t_total = time.perf_counter()
    _perf_logger.info(f"[PERF] extract_tables_fast TOTAL: {t_total-t0:.2f}s")
    return tables
